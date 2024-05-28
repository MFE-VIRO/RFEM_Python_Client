import os
import sys
baseName = os.path.basename(__file__)
dirName = os.path.dirname(__file__)
print('basename:    ', baseName)
print('dirname:     ', dirName)
sys.path.append(dirName + r'/../..')

import pandas as pd
import xlwings as xw
import openpyxl as xl

# from RFEM.enums import *
from RFEM.initModel import Model
from RFEM.enums import ObjectTypes
from RFEM.LoadCasesAndCombinations.loadCombination import LoadCombination
from RFEM.Tools.GetObjectNumbersByType import GetObjectNumbersByType

if __name__ == '__main__':
    Model(False, "") #Work in current RFEM6 model

    # Model.clientModel.service.begin_modification()
    lc_numbers = GetObjectNumbersByType(ObjectType=ObjectTypes.E_OBJECT_TYPE_LOAD_CASE)
    co_numbers = GetObjectNumbersByType(ObjectType=ObjectTypes.E_OBJECT_TYPE_LOAD_COMBINATION)
    co_items = {}
    for c in co_numbers:
        co = Model.clientModel.service.get_load_combination(c)
        items = co.items.load_combination_items
        dictItems_c = {}
        for i in range(len(items)):
            dictItems_c[items[i].row.load_case]=items[i].row.factor
        co_items[c]=dictItems_c

    print(co_items[21][300])

    try:
        wb = xl.load_workbook('MFE_Combinaties.xlsx')
    except:
        wb = xl.Workbook('MFE_Combinaties.xlsx')
    try:
        wb["MFE_Combinaties"]
    except:
        wb.create_sheet("MFE_Combinaties")

    i=2
    for l in lc_numbers:
        wb["MFE_Combinaties"].cell(1,i,l)
        i=i+1
    wb.save('MFE_Combinaties.xlsx')
    # pd.ExcelWriter('PandasEersteExcel.xlsm')
    # pd.ExcelFile('PandasEersteExcel.xlsm')
    print()

    # Model.clientModel.service.finish_modification()


